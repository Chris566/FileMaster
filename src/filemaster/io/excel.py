"""Excel 报告生成（W4 详细实现）.

基于 openpyxl：多 sheet、冻结表头、自动筛选、列宽自适应。
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ReportRow:
    """报告单行."""

    original_name: str
    new_name: str
    category: str
    status: str  # "OK" / "SKIPPED" / "CONFLICT" / "ERROR" / "DRY_RUN"
    copy_status: str  # "OK" / "SKIPPED" / "OVERWRITTEN" / ""
    source_path: str
    target_path: str
    mode: str = ""  # "NORMAL" / "DRY-RUN"


class ExcelReporter:
    """Excel 报告生成器."""

    HEADERS = (
        "原文件名",
        "新文件名",
        "文件类型",
        "状态",
        "复制状态",
        "原路径",
        "目标路径",
        "模式",
    )

    def __init__(self, source_dir_name: str = "report") -> None:
        self._source_dir_name = source_dir_name

    def export(
        self,
        rows: Iterable[ReportRow],
        output_path: Path,
        *,
        title: str = "FileMaster 处理报告",
    ) -> Path:
        """导出 Excel.

        Args:
            rows: 数据行
            output_path: 输出路径
            title: 报表标题（写 A1）
        Returns:
            实际写入路径
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "处理明细"

        # 标题行
        ws.cell(row=1, column=1, value=title)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="0078D4")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.HEADERS))

        # 表头
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0078D4")
            cell.alignment = Alignment(horizontal="center")

        # 数据
        row_idx = 3
        for r in rows:
            for col_idx, value in enumerate(
                (
                    r.original_name,
                    r.new_name,
                    r.category,
                    r.status,
                    r.copy_status,
                    r.source_path,
                    r.target_path,
                    r.mode,
                ),
                1,
            ):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        # 冻结表头
        ws.freeze_panes = "A3"

        # 自动筛选
        if row_idx > 3:
            ws.auto_filter.ref = f"A2:H{row_idx - 1}"

        # 列宽自适应（最长字符串 + 2）
        for col_idx in range(1, len(self.HEADERS) + 1):
            letter = get_column_letter(col_idx)
            max_len = max(
                [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, row_idx)]
                + [len(self.HEADERS[col_idx - 1])]
            )
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
