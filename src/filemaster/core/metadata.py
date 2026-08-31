"""元数据读取（W3 + W5 扩展）.

- PDF：PyMuPDF (fitz)
- Word：python-docx
- Excel：openpyxl
- 图片：Pillow EXIF

W5 新增字段（按文件类型命名空间）：
- PDF: pdf_pages (page_count 别名) / pdf_created / pdf_modified
- Word: word_paragraphs / word_created / word_modified
- Excel: excel_sheets (sheets count) / excel_created / excel_modified
- Image: image_taken_at (EXIF DateTimeOriginal) / image_camera_make / image_camera_model
        / image_format / image_aspect_ratio (e.g. "16:9")
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class FileMetadata:
    """文件元数据.

    通用字段（W3）+ W5 扩展字段。
    命名空间字段在对应文件类型时填充；非对应文件类型保持默认值。
    """

    # ---- 通用 (W3) ----
    title: str = ""
    author: str = ""
    subject: str = ""
    created: str = ""  # ISO 8601 或 EXIF DateTime
    modified: str = ""
    page_count: int = 0
    # ---- Word (W5) ----
    paragraphs: int = 0
    # ---- Excel (W5) ----
    sheets_count: int = 0
    # ---- Image (W5) ----
    taken_at: str = ""  # EXIF DateTimeOriginal
    camera_make: str = ""
    camera_model: str = ""
    image_format: str = ""
    width: int = 0
    height: int = 0
    aspect_ratio: str = ""  # e.g. "16:9" / "4:3" / "1:1"
    # ---- 原始 extra ----
    extra: dict = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.extra is None:
            object.__setattr__(self, "extra", {})


def _compute_aspect_ratio(width: int, height: int) -> str:
    """从宽高算最接近的常用纵横比字符串. 简化版, 不算 GCD."""
    if width <= 0 or height <= 0:
        return ""
    # 常用比例
    ratio = width / height
    for w, h, name in [
        (16, 9, "16:9"),
        (4, 3, "4:3"),
        (3, 2, "3:2"),
        (1, 1, "1:1"),
        (21, 9, "21:9"),
        (5, 4, "5:4"),
        (2, 3, "2:3"),
        (9, 16, "9:16"),
    ]:
        if abs(ratio - w / h) < 0.02:  # 容差 2%
            return name
    # 兜底: 真实比例 (e.g. "1.85:1")
    return f"{ratio:.2f}:1"


class MetadataReader:
    """元数据读取器（按扩展名分发）."""

    PDF_EXTS = {".pdf"}
    WORD_EXTS = {".doc", ".docx", ".rtf"}
    EXCEL_EXTS = {".xls", ".xlsx", ".ods"}
    IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tif", ".tiff"}

    def read(self, file: Path) -> FileMetadata:
        """读取文件元数据.

        Args:
            file: 源文件
        Returns:
            FileMetadata
        """
        ext = file.suffix.lower()
        if ext in self.PDF_EXTS:
            return self._read_pdf(file)
        if ext in self.WORD_EXTS:
            return self._read_word(file)
        if ext in self.EXCEL_EXTS:
            return self._read_excel(file)
        if ext in self.IMAGE_EXTS:
            return self._read_image(file)
        return FileMetadata()

    def _read_pdf(self, file: Path) -> FileMetadata:
        """PDF：PyMuPDF."""
        try:
            import fitz  # PyMuPDF

            with fitz.open(file) as doc:
                meta = doc.metadata or {}
                return FileMetadata(
                    title=meta.get("title", "") or "",
                    author=meta.get("author", "") or "",
                    subject=meta.get("subject", "") or "",
                    created=str(meta.get("creationDate", "") or ""),
                    modified=str(meta.get("modDate", "") or ""),
                    page_count=doc.page_count,
                )
        except Exception:
            return FileMetadata()

    def _read_word(self, file: Path) -> FileMetadata:
        """Word：python-docx."""
        try:
            from docx import Document

            doc = Document(file)
            cp = doc.core_properties
            paragraphs = len(doc.paragraphs)
            # 简化版 page_count ≈ paragraph 数 (Word 真分页需渲染器)
            page_count = max(paragraphs, 1) if paragraphs > 0 else 0
            return FileMetadata(
                title=cp.title or "",
                author=cp.author or "",
                subject=cp.subject or "",
                created=str(cp.created or ""),
                modified=str(cp.modified or ""),
                page_count=page_count,
                paragraphs=paragraphs,
            )
        except Exception:
            return FileMetadata()

    def _read_excel(self, file: Path) -> FileMetadata:
        """Excel：openpyxl."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file, read_only=True, data_only=True)
            props = wb.properties
            sheets_count = len(wb.sheetnames)
            return FileMetadata(
                title=props.title or "",
                author=props.creator or "",
                subject=props.subject or "",
                created=str(props.created or ""),
                modified=str(props.modified or ""),
                sheets_count=sheets_count,
            )
        except Exception:
            return FileMetadata()

    def _read_image(self, file: Path) -> FileMetadata:
        """图片：Pillow EXIF."""
        try:
            from PIL import Image
            from PIL.ExifTags import TAGS

            with Image.open(file) as img:
                exif_raw = img.getexif() or {}
                decoded = {TAGS.get(k, k): v for k, v in exif_raw.items()}
                width, height = img.size
                # DateTimeOriginal 在 ExifTags 子表里 (36867), 主表只有 DateTime (306)
                taken_at = (
                    decoded.get("DateTimeOriginal", "")
                    or decoded.get("DateTime", "")
                    or ""
                )
                return FileMetadata(
                    title="",
                    author=decoded.get("Artist", "") or "",
                    created=str(decoded.get("DateTime", "") or ""),
                    taken_at=str(taken_at or ""),
                    camera_make=str(decoded.get("Make", "") or ""),
                    camera_model=str(decoded.get("Model", "") or ""),
                    image_format=img.format or "",
                    width=width,
                    height=height,
                    aspect_ratio=_compute_aspect_ratio(width, height),
                    extra={"size": img.size, "format": img.format},
                )
        except Exception:
            return FileMetadata()

    def read_all(self, files: Iterable[Path]) -> dict[Path, FileMetadata]:
        """批量读取."""
        return {f: self.read(f) for f in files}
